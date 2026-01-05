# -*- coding: utf-8 -*-
"""
CSV 파일에서 내용 안의 모든 세미콜론을 쉼표로 변경
"""
import sys
import os
import re

# Windows 콘솔 인코딩 설정
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

def fix_all_semicolons(csv_file):
    """내용 필드의 모든 세미콜론을 쉼표로 변경"""
    with open(csv_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    lines = content.split('\n')
    fixed_lines = []
    changes = []
    
    for i, line in enumerate(lines, 1):
        original = line
        
        # 헤더는 그대로
        if i == 1:
            fixed_lines.append(line)
            continue
        
        if not line.strip():
            fixed_lines.append(line)
            continue
        
        # 세미콜론으로 분리 (최대 6개 필드)
        parts = line.split(';')
        
        if len(parts) < 4:
            fixed_lines.append(line)
            continue
        
        차시 = parts[0]
        유형 = parts[1]
        번호 = parts[2]
        
        # 내용 필드: 3번째 인덱스부터 정답 필드 전까지
        # 정답 필드는 마지막에서 두 번째 또는 세 번째
        # 해설 필드는 마지막
        
        # 내용 필드 찾기 (3번째 인덱스)
        내용 = parts[3]
        
        # 정답 필드 찾기 (마지막에서 두 번째 또는 세 번째)
        if len(parts) >= 5:
            # 정답과 해설이 있는 경우
            if len(parts) == 6:
                정답 = parts[4]
                해설 = parts[5]
            else:
                # 정답만 있는 경우
                정답 = parts[4] if len(parts) > 4 else ''
                해설 = ''
        else:
            정답 = ''
            해설 = ''
        
        # 내용 필드의 세미콜론을 쉼표로 변경
        original_content = 내용
        # 괄호 안의 세미콜론 처리: (KRW; USD; EUR) -> (KRW, USD, EUR)
        def replace_in_parens(match):
            text = match.group(1)
            # 세미콜론을 쉼표로 변경
            text = text.replace(';', ',')
            return '(' + text + ')'
        내용 = re.sub(r'\(([^)]+)\)', replace_in_parens, 내용)
        # 일반적인 세미콜론을 쉼표로 (공백이 있는 경우)
        내용 = 내용.replace('; ', ', ')
        # 나머지 세미콜론도 쉼표로
        if ';' in 내용:
            내용 = 내용.replace(';', ',')
        
        if 내용 != original_content:
            changes.append(f"줄 {i}: 내용")
        
        # 해설 필드의 세미콜론을 쉼표로 변경
        original_explanation = 해설
        해설 = re.sub(r'\(([^)]+)\)', replace_in_parens, 해설)
        해설 = 해설.replace('; ', ', ')
        if ';' in 해설:
            해설 = 해설.replace(';', ',')
        
        if 해설 != original_explanation:
            changes.append(f"줄 {i}: 해설")
        
        # 수정된 라인 재구성
        if 해설:
            new_line = f"{차시};{유형};{번호};{내용};{정답};{해설}\n"
        else:
            new_line = f"{차시};{유형};{번호};{내용};{정답};\n"
        fixed_lines.append(new_line)
    
    # 저장
    with open(csv_file, 'w', encoding='utf-8', newline='') as f:
        f.writelines(fixed_lines)
    
    return changes

def main():
    csv_file = '각_차시별_학습정리.csv'
    
    if not os.path.exists(csv_file):
        print(f'[오류] 파일을 찾을 수 없습니다: {csv_file}')
        return
    
    print(f'[진행] 세미콜론 -> 쉼표 변경 중: {csv_file}')
    changes = fix_all_semicolons(csv_file)
    
    print(f'[성공] {len(changes)}개 수정 완료')
    if changes:
        for change in changes[:15]:
            print(f'  - {change}')
        if len(changes) > 15:
            print(f'  ... 외 {len(changes) - 15}개')
    
    # 엑셀 파일 재생성
    print('\n[진행] 엑셀 파일 재생성 중...')
    try:
        import pandas as pd
        from collections import defaultdict
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # CSV 읽기
        df = pd.read_csv(csv_file, sep=';', encoding='utf-8', on_bad_lines='skip', engine='python')
        
        # 차시별 그룹화
        grouped = defaultdict(list)
        for idx, row in df.iterrows():
            차시 = row['차시']
            grouped[차시].append(row.to_dict())
        
        excel_file = csv_file.replace('.csv', '_포맷팅.xlsx')
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            sorted_차시 = sorted(grouped.keys(), key=lambda x: int(x.replace('차시', '')) if x.replace('차시', '').isdigit() else 999)
            
            for 차시 in sorted_차시:
                차시_데이터 = grouped[차시]
                
                퀴즈 = [d for d in 차시_데이터 if d['유형'] == '퀴즈']
                학습목표 = [d for d in 차시_데이터 if d['유형'] == '학습목표']
                학습정리 = [d for d in 차시_데이터 if d['유형'] == '학습정리']
                
                formatted_rows = []
                
                # Quiz 섹션
                if 퀴즈:
                    formatted_rows.append({
                        'A': 'Quiz',
                        'B': '',
                        'C': '',
                        'D': '정답',
                        'E': '해설'
                    })
                    
                    for item in 퀴즈:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': item['정답'] if pd.notna(item['정답']) else '',
                            'E': item['해설'] if pd.notna(item['해설']) else ''
                        })
                    
                    formatted_rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': ''})
                
                # 학습목표 섹션
                if 학습목표:
                    formatted_rows.append({
                        'A': '학습목표',
                        'B': '',
                        'C': '',
                        'D': '',
                        'E': ''
                    })
                    
                    for item in 학습목표:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': '',
                            'E': ''
                        })
                    
                    formatted_rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': ''})
                
                # 학습정리 섹션
                if 학습정리:
                    formatted_rows.append({
                        'A': '학습정리',
                        'B': '',
                        'C': '',
                        'D': '',
                        'E': ''
                    })
                    
                    for item in 학습정리:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': '',
                            'E': ''
                        })
                
                if formatted_rows:
                    formatted_df = pd.DataFrame(formatted_rows)
                    sheet_name = 차시[:31] if len(차시) > 31 else 차시
                    formatted_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        # 포맷팅 적용
        wb = load_workbook(excel_file)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        bold_font = Font(bold=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 80
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 50
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                cell_a = row[0] if len(row) > 0 else None
                cell_b = row[1] if len(row) > 1 else None
                cell_c = row[2] if len(row) > 2 else None
                cell_d = row[3] if len(row) > 3 else None
                cell_e = row[4] if len(row) > 4 else None
                
                if cell_a and cell_a.value in ['Quiz', '학습목표', '학습정리']:
                    cell_a.fill = header_fill
                    cell_a.font = bold_font
                    if cell_b:
                        cell_b.fill = header_fill
                        cell_b.font = bold_font
                    if cell_a.value == 'Quiz':
                        if cell_d:
                            cell_d.fill = header_fill
                            cell_d.font = bold_font
                        if cell_e:
                            cell_e.fill = header_fill
                            cell_e.font = bold_font
                
                if cell_b and cell_b.value and str(cell_b.value).isdigit():
                    if cell_b:
                        cell_b.alignment = Alignment(horizontal='center', vertical='top')
                    if cell_c:
                        cell_c.alignment = Alignment(wrap_text=True, vertical='top')
                    if cell_d:
                        cell_d.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                    if cell_e:
                        cell_e.alignment = Alignment(wrap_text=True, vertical='top')
        
        wb.save(excel_file)
        print(f'[성공] 엑셀 파일 생성 완료: {excel_file}')
        print(f'       총 {len(grouped)}개 차시 시트 생성됨')
        
    except Exception as e:
        print(f'[오류] 엑셀 파일 생성 실패: {e}')
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()

