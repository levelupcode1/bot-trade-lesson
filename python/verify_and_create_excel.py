# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys

def verify_csv(csv_file):
    """CSV 파일 검증"""
    print(f"[검증] CSV 파일 읽기: {csv_file}")
    try:
        df = pd.read_csv(csv_file, sep=';', encoding='utf-8-sig')
        print(f"[검증] 총 행 수: {len(df)}")
        
        # 퀴즈 통계
        quiz_df = df[df['유형'] == '퀴즈']
        o_count = len(quiz_df[quiz_df['정답'] == 'O'])
        x_count = len(quiz_df[quiz_df['정답'] == 'X'])
        print(f"[검증] 퀴즈 O 개수: {o_count}")
        print(f"[검증] 퀴즈 X 개수: {x_count}")
        
        # 차시별 퀴즈 개수
        print("\n[검증] 차시별 퀴즈 개수:")
        quiz_by_chapter = quiz_df.groupby('차시').size()
        for chapter, count in quiz_by_chapter.items():
            chapter_quiz = quiz_df[quiz_df['차시'] == chapter]
            chapter_o = len(chapter_quiz[chapter_quiz['정답'] == 'O'])
            chapter_x = len(chapter_quiz[chapter_quiz['정답'] == 'X'])
            print(f"  {chapter}: 총 {count}개 (O: {chapter_o}, X: {chapter_x})")
        
        # 세미콜론 검사
        print("\n[검증] 세미콜론 검사 중...")
        semicolon_count = 0
        for idx, row in df.iterrows():
            content = str(row['내용'])
            if ';' in content and '차시;유형' not in content:
                # 필드 구분자가 아닌 내용 내 세미콜론 검사
                if not content.startswith('차시'):
                    semicolon_count += 1
                    if semicolon_count <= 5:  # 처음 5개만 출력
                        print(f"  [경고] 행 {idx+2}: 내용에 세미콜론 발견")
                        print(f"         내용: {content[:50]}...")
        
        if semicolon_count > 0:
            print(f"  [경고] 총 {semicolon_count}개 행에서 세미콜론 발견")
        else:
            print("  [성공] 내용 내 세미콜론 없음")
        
        return df
    except Exception as e:
        print(f"[오류] CSV 파일 읽기 실패: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_formatted_excel(csv_file, excel_file='각_차시별_학습정리_포맷팅.xlsx'):
    """포맷팅된 엑셀 파일 생성"""
    print(f"\n[생성] 엑셀 파일 생성 시작: {excel_file}")
    
    try:
        df = pd.read_csv(csv_file, sep=';', encoding='utf-8-sig')
        
        # 차시별로 그룹화
        grouped = {}
        for _, row in df.iterrows():
            차시 = row['차시']
            if 차시 not in grouped:
                grouped[차시] = []
            grouped[차시].append(row.to_dict())
        
        # 차시 정렬
        sorted_차시 = sorted(grouped.keys(), key=lambda x: int(x.replace('차시', '')) if x.replace('차시', '').isdigit() else 999)
        
        # 엑셀 파일 생성
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            for 차시 in sorted_차시:
                차시_데이터 = grouped[차시]
                
                # 유형별로 분류
                퀴즈 = [d for d in 차시_데이터 if d['유형'] == '퀴즈']
                학습목표 = [d for d in 차시_데이터 if d['유형'] == '학습목표']
                학습정리 = [d for d in 차시_데이터 if d['유형'] == '학습정리']
                
                formatted_rows = []
                
                # Quiz 섹션
                if 퀴즈:
                    formatted_rows.append({'A': 'Quiz', 'B': '', 'C': '', 'D': '', 'E': ''})
                    for item in 퀴즈:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': item['정답'] if pd.notna(item['정답']) else '',
                            'E': item['해설'] if pd.notna(item['해설']) else ''
                        })
                    formatted_rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': ''})
                
                # 학습목표 섹션
                if 학습목표:
                    formatted_rows.append({'A': '학습목표', 'B': '', 'C': '', 'D': '', 'E': ''})
                    for item in 학습목표:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': '',
                            'E': ''
                        })
                    formatted_rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': ''})
                
                # 학습정리 섹션
                if 학습정리:
                    formatted_rows.append({'A': '학습정리', 'B': '', 'C': '', 'D': '', 'E': ''})
                    for item in 학습정리:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': '',
                            'E': ''
                        })
                
                formatted_df = pd.DataFrame(formatted_rows)
                sheet_name = 차시[:31] if len(차시) > 31 else 차시
                formatted_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                # 포맷팅 적용
                workbook = writer.book
                sheet = workbook[sheet_name]
                
                # 열 너비 설정
                sheet.column_dimensions['A'].width = 15
                sheet.column_dimensions['B'].width = 5
                sheet.column_dimensions['C'].width = 70
                sheet.column_dimensions['D'].width = 10
                sheet.column_dimensions['E'].width = 50
                
                # 스타일 적용
                header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                
                for r_idx, row in enumerate(sheet.iter_rows(), 1):
                    if row[0].value in ['Quiz', '학습목표', '학습정리']:
                        for cell in row:
                            cell.fill = header_fill
                            cell.font = Font(bold=True)
                        sheet.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=5)
                        row[0].alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"[성공] 엑셀 파일 생성 완료: {excel_file}")
        print(f"       총 {len(sorted_차시)}개 차시 시트 생성됨")
        return True
        
    except Exception as e:
        print(f"[오류] 엑셀 파일 생성 실패: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    csv_file = '각_차시별_학습정리.csv'
    
    # CSV 검증
    df = verify_csv(csv_file)
    
    if df is not None:
        # 엑셀 파일 생성
        create_formatted_excel(csv_file)
    else:
        print("[오류] CSV 파일 검증 실패로 엑셀 파일을 생성할 수 없습니다.")
        sys.exit(1)

