# -*- coding: utf-8 -*-
"""
prompts 폴더의 프롬프트 파일들을 엑셀 파일로 변환하는 스크립트
"""
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_prompts_from_markdown(file_path):
    """마크다운 파일에서 프롬프트 추출"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    prompts = []
    lines = content.split('\n')
    
    i = 0
    while i < len(lines):
        line = lines[i]
        
        # 프롬프트 번호 찾기
        match = re.match(r'^## (\d+)번 프롬프트', line)
        if match:
            prompt_num = match.group(1)
            prompt_content = []
            
            # 다음 줄부터 프롬프트 내용 찾기
            i += 1
            while i < len(lines):
                current_line = lines[i]
                
                # ```text 시작
                if current_line.strip() == '```text':
                    i += 1
                    # ``` 까지 읽기
                    while i < len(lines):
                        if lines[i].strip() == '```':
                            break
                        prompt_content.append(lines[i])
                        i += 1
                    break
                elif current_line.strip().startswith('```'):
                    break
                elif re.match(r'^---', current_line):
                    break
                i += 1
            
            if prompt_content:
                prompts.append({
                    'num': prompt_num,
                    'content': '\n'.join(prompt_content).strip()
                })
        else:
            i += 1
    
    return prompts

def create_excel_file(prompts_dir, output_file):
    """프롬프트들을 엑셀 파일로 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "프롬프트 모음"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 작성
    headers = ['차시', '프롬프트 번호', '프롬프트 내용']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 데이터 작성
    row_num = 2
    lesson_files = sorted([f for f in prompts_dir.glob('lesson-*.md')])
    
    for lesson_file in lesson_files:
        # 차시 번호 추출
        match = re.search(r'lesson-(\d+)-prompts', lesson_file.name)
        if not match:
            continue
        
        lesson_num = int(match.group(1))
        prompts = extract_prompts_from_markdown(lesson_file)
        
        for prompt in prompts:
            # 차시
            cell1 = ws.cell(row=row_num, column=1, value=f"{lesson_num}차시")
            cell1.alignment = Alignment(horizontal='center', vertical='top')
            cell1.border = border
            cell1.font = Font(bold=True)
            
            # 프롬프트 번호
            cell2 = ws.cell(row=row_num, column=2, value=f"{prompt['num']}번")
            cell2.alignment = Alignment(horizontal='center', vertical='top')
            cell2.border = border
            
            # 프롬프트 내용
            cell3 = ws.cell(row=row_num, column=3, value=prompt['content'])
            cell3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell3.border = border
            
            row_num += 1
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 80
    
    # 행 높이 조정 (자동)
    for row in ws.iter_rows(min_row=2):
        max_height = 0
        for cell in row:
            if cell.value:
                lines = str(cell.value).split('\n')
                max_height = max(max_height, len(lines))
        if max_height > 1:
            ws.row_dimensions[row[0].row].height = min(max_height * 15, 200)
    
    # 첫 번째 행 높이
    ws.row_dimensions[1].height = 25
    
    # 저장
    wb.save(output_file)
    print(f"✓ 엑셀 파일 생성 완료: {output_file}")
    print(f"  총 {row_num - 2}개 프롬프트 포함")

def main():
    base_dir = Path(__file__).parent.parent
    prompts_dir = base_dir / 'prompts'
    output_file = base_dir / '프롬프트_모음.xlsx'
    
    if not prompts_dir.exists():
        print(f"❌ prompts 폴더를 찾을 수 없습니다: {prompts_dir}")
        return
    
    create_excel_file(prompts_dir, output_file)

if __name__ == '__main__':
    main()
