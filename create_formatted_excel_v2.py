# -*- coding: utf-8 -*-
"""
CSV 파일을 읽어서 각 차시별로 포맷된 엑셀 시트를 만드는 스크립트 (수정 버전)
- 차시 컬럼 제거
- 문항 번호 컬럼 제거
- 섹션 헤더 밑에 번호와 내용이 바로 나오도록
- 정답과 해설 분리
- Column C 노란색 배경/더블 보더 제거
"""
import pandas as pd
import sys
import os
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Windows 콘솔 인코딩 설정
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

def create_formatted_excel(csv_file, excel_file=None):
    """
    CSV 파일을 읽어서 각 차시별로 포맷된 엑셀 시트를 생성
    
    구조:
    - Column A: 섹션 헤더 (Quiz, 학습목표, 학습정리)
    - Column B: 번호
    - Column C: 내용 (퀴즈면 문제, 학습목표면 학습목표 내용, 학습정리면 학습정리 내용)
    - Column D: 정답 (퀴즈만)
    - Column E: 해설 (퀴즈만)
    """
    try:
        # CSV 파일 읽기
        print(f'[진행] CSV 파일 읽는 중: {csv_file}')
        df = pd.read_csv(csv_file, sep=';', encoding='utf-8', on_bad_lines='skip', engine='python')
        print(f'[정보] 총 {len(df)}개 행 읽기 완료')
        
        # 차시별로 그룹화
        grouped = defaultdict(list)
        for idx, row in df.iterrows():
            차시 = row['차시']
            grouped[차시].append(row.to_dict())
        
        # 엑셀 파일명이 없으면 자동 생성
        if excel_file is None:
            excel_file = csv_file.replace('.csv', '_포맷팅.xlsx')
        
        # 엑셀 파일 생성
        print(f'[진행] 엑셀 파일 생성 중: {excel_file}')
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 차시별로 정렬하여 시트 생성
            sorted_차시 = sorted(grouped.keys(), key=lambda x: int(x.replace('차시', '')) if x.replace('차시', '').isdigit() else 999)
            
            for 차시 in sorted_차시:
                차시_데이터 = grouped[차시]
                
                # 유형별로 분류
                퀴즈 = [d for d in 차시_데이터 if d['유형'] == '퀴즈']
                학습목표 = [d for d in 차시_데이터 if d['유형'] == '학습목표']
                학습정리 = [d for d in 차시_데이터 if d['유형'] == '학습정리']
                
                # 새로운 데이터프레임 생성
                formatted_rows = []
                
                # Quiz 섹션
                if 퀴즈:
                    formatted_rows.append({
                        'A': 'Quiz',
                        'B': '',
                        'C': '',
                        'D': '정답',
                        'E': '해설'
                    })
                    
                    for item in 퀴즈:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': item['정답'] if pd.notna(item['정답']) else '',
                            'E': item['해설'] if pd.notna(item['해설']) else ''
                        })
                    
                    # 빈 행 추가
                    formatted_rows.append({
                        'A': '',
                        'B': '',
                        'C': '',
                        'D': '',
                        'E': ''
                    })
                
                # 학습목표 섹션
                if 학습목표:
                    formatted_rows.append({
                        'A': '학습목표',
                        'B': '',
                        'C': '',
                        'D': '',
                        'E': ''
                    })
                    
                    for item in 학습목표:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': '',
                            'E': ''
                        })
                    
                    # 빈 행 추가
                    formatted_rows.append({
                        'A': '',
                        'B': '',
                        'C': '',
                        'D': '',
                        'E': ''
                    })
                
                # 학습정리 섹션
                if 학습정리:
                    formatted_rows.append({
                        'A': '학습정리',
                        'B': '',
                        'C': '',
                        'D': '',
                        'E': ''
                    })
                    
                    for item in 학습정리:
                        formatted_rows.append({
                            'A': '',
                            'B': str(item['번호']),
                            'C': item['내용'],
                            'D': '',
                            'E': ''
                        })
                
                # 데이터프레임 생성
                formatted_df = pd.DataFrame(formatted_rows)
                
                # 시트 이름은 최대 31자
                sheet_name = 차시[:31] if len(차시) > 31 else 차시
                formatted_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        # 포맷팅 적용
        print(f'[진행] 포맷팅 적용 중...')
        wb = load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 스타일 정의
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # 회색
            bold_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 컬럼 너비 설정
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 80
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 50
            
            # 각 행 포맷팅
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
                cell_a = row[0]  # Column A
                cell_b = row[1] if len(row) > 1 else None  # Column B
                cell_c = row[2] if len(row) > 2 else None  # Column C
                cell_d = row[3] if len(row) > 3 else None  # Column D
                cell_e = row[4] if len(row) > 4 else None  # Column E
                
                # 섹션 헤더 (Quiz, 학습목표, 학습정리)
                if cell_a.value in ['Quiz', '학습목표', '학습정리']:
                    cell_a.fill = header_fill
                    cell_a.font = bold_font
                    # Quiz 섹션의 경우 D, E도 헤더 스타일
                    if cell_a.value == 'Quiz':
                        if cell_d:
                            cell_d.fill = header_fill
                            cell_d.font = bold_font
                        if cell_e:
                            cell_e.fill = header_fill
                            cell_e.font = bold_font
                
                # 번호가 있는 행 (데이터 행)
                elif cell_b and cell_b.value and str(cell_b.value).isdigit():
                    # Column B (번호) - 중앙 정렬
                    if cell_b:
                        cell_b.alignment = Alignment(horizontal='center', vertical='top')
                    
                    # Column C (내용) - 텍스트 래핑, 상단 정렬 (노란색 배경/더블 보더 제거)
                    if cell_c:
                        cell_c.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Column D, E (정답/해설) - 텍스트 래핑, 상단 정렬
                    if cell_d:
                        cell_d.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                    if cell_e:
                        cell_e.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 행 높이 자동 조정
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        ws.row_dimensions[cell.row].height = None
        
        wb.save(excel_file)
        print(f'\n[성공] 포맷팅된 엑셀 파일 생성 완료: {excel_file}')
        print(f'       총 {len(grouped)}개 차시 시트 생성됨')
        
        return excel_file
        
    except ImportError:
        print('[오류] openpyxl이 설치되지 않았습니다.')
        print('       다음 명령을 실행하세요: pip install openpyxl')
        return None
    except Exception as e:
        print(f'[오류] 변환 중 오류 발생: {e}')
        import traceback
        traceback.print_exc()
        return None

def main():
    csv_file = '각_차시별_학습정리.csv'
    
    if not os.path.exists(csv_file):
        print(f'[오류] 파일을 찾을 수 없습니다: {csv_file}')
        return
    
    create_formatted_excel(csv_file)

if __name__ == '__main__':
    main()

