#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 리포트 생성기
openpyxl을 사용한 Excel 생성
"""

from .base_generator import BaseReportGenerator
from typing import Dict, Any
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExcelReportGenerator(BaseReportGenerator):
    """Excel 리포트 생성 클래스"""
    
    def generate(self, report_type: str, data: Dict[str, Any]) -> str:
        """Excel 리포트 생성"""
        try:
            # openpyxl 라이브러리 확인
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                logger.warning("openpyxl이 설치되지 않았습니다. Excel 생성을 건너뜁니다.")
                return ""
            
            filepath = self._get_output_filename(report_type, 'xlsx')
            
            wb = Workbook()
            ws = wb.active
            ws.title = "요약"
            
            # 헤더 스타일
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            
            # 제목
            title_map = {
                'daily': '일간 거래 리포트',
                'weekly': '주간 성과 리포트',
                'monthly': '월간 종합 리포트',
                'alert': '긴급 알림 리포트'
            }
            
            ws['A1'] = title_map.get(report_type, '거래 리포트')
            ws['A1'].font = Font(bold=True, size=18)
            
            # 핵심 지표
            analysis = data.get('analysis', {})
            metrics = [
                ('총 수익률', f"{analysis.get('total_return', 0):.2f}%"),
                ('거래 수', analysis.get('total_trades', 0)),
                ('승률', f"{analysis.get('win_rate', 0):.1f}%"),
                ('최대 낙폭', f"{analysis.get('max_drawdown', 0):.2f}%"),
                ('샤프 비율', f"{analysis.get('sharpe_ratio', 0):.2f}"),
            ]
            
            row = 3
            for metric, value in metrics:
                ws.cell(row, 1, metric).font = Font(bold=True)
                ws.cell(row, 2, value)
                row += 1
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            
            # 저장
            wb.save(filepath)
            
            logger.info(f"Excel 리포트 생성: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Excel 생성 오류: {e}", exc_info=True)
            return ""

